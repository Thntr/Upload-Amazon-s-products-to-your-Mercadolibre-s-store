import json
from openpyxl import load_workbook

def get_json(json_file, path):

    incomplete_path = path
    theJson = "\\" + json_file
    path_of_definitiveList = str(incomplete_path + theJson).replace("\\", '/')

    return path_of_definitiveList

activation_of_external_scripts = False

if activation_of_external_scripts == True:

    incomplete_path = r"C:\Users\ramzhacker\Documents\MEGA\Newy\Wink_Inc\WinkAI\#WinkAIMLApp\pythonSdkMaster\winktech\project\winkscraper\winkscraper\spiders\xlsxDataItems"
    xlsx_file = "anotheramzfckr.xlsx"
    theXlsx = "\\" + xlsx_file
    path_of_definitiveList = str(incomplete_path + theXlsx).replace("\\", '/')

    workbook = load_workbook(filename=path_of_definitiveList)

    i = 2
    arrayoflinks = []

    while i <= 9001:
        if workbook.worksheets[0].cell(row=i, column=4).value != None:
            arrayoflinks.append(workbook.worksheets[0].cell(row=i, column=4).value)

        i += 1


    print(len(arrayoflinks))

    path = r"C:\Users\ramzhacker\Documents\MEGA\Newy\Wink_Inc\WinkAI\#WinkAIMLApp\pythonSdkMaster\winktech\project\winkscraper\winkscraper\spiders\items"
    json_target = get_json("toysItems.json", path=path)

    with open(json_target, "r") as jsonFile:
        data = json.load(jsonFile)

    tmp = data["startUrl"]
    data["startUrl"] = arrayoflinks

    print(len(data["startUrl"]))

    with open(json_target, "w") as jsonFile:
        json.dump(data, jsonFile)