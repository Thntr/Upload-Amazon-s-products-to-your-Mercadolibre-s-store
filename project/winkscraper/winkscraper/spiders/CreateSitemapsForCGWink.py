import json
from openpyxl import load_workbook

activation_of_external_scripts = False

if activation_of_external_scripts == True:

    workbook = load_workbook(filename="cgWinkItems.xlsx")

    i = 2
    arrayoflinks = []

    while i <= 1087:
        if workbook.worksheets[0].cell(row=i, column=4).value != None:
            arrayoflinks.append(workbook.worksheets[0].cell(row=i, column=4).value)

        i += 1


    print(len(arrayoflinks))

    with open("cgDefinitiveList.json", "r") as jsonFile:
        data = json.load(jsonFile)

    tmp = data["startUrl"]
    data["startUrl"] = arrayoflinks

    print(len(data["startUrl"]))

    with open("cgDefinitiveList.json", "w") as jsonFile:
        json.dump(data, jsonFile)
